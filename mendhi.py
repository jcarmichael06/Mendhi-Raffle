from openpyxl import load_workbook
from operator import itemgetter
from tkinter import messagebox, filedialog
import random

#importing zipcodes from excel file
zips = {}

wb = load_workbook('ZipCodes.xlsx')
ws = wb.active
for r in range(1,101):
	for c in range(1,11):
		a = str(ws.cell(row=r,column=c).value).split(' ',1)
		a[1] = a[1].split('\n')
		a[1] = a[1][0] + ' ' + a[1][1]
		zips[a[0]] = a[1]


#raffle code
filename = filedialog.askopenfilename(filetypes=[("Excel file","*.xlsx")])
wb = load_workbook(filename)
for sheet in wb:
	if sheet.title != 'Sheet1':
		wb.remove(sheet)
contestantSheet = wb.create_sheet("Valid Contestants")
cheatSheet = wb.create_sheet("Cheaters")
ws = wb["Sheet1"]

class Contestants:

	def __init__(self):
		self.contestants = []
		self.cheaters = []
		self.suspiciousPeeps = []
		self.createContestants()
		self.findSuspiciousPeople()
		self.deleteSuspiciousPerson()
		self.publishCheaters()
		self.publishContestants()

	def createContestants(self):
		for r in range(1,ws.max_row+1):
			person = []
			person.append(str(ws.cell(row=r,column=2).value).upper().strip().split(' ')[0])
			person.append(str(ws.cell(row=r,column=3).value).upper().strip().split(' ')[0])
			person.append(str(ws.cell(row=r,column=4).value).strip())
			person.append(str(int(ws.cell(row=r,column=5).value)))
			person.append(str(ws.cell(row=r,column=6).value).upper().strip().split(' ')[0])
			if person in self.cheaters:
				pass
			elif person in self.contestants:
				self.cheaters.append(person)
				self.contestants.remove(person)
			else:
				self.contestants.append(person)
	
	def publishCheaters(self):
		for x in range(len(self.cheaters)):
			cheatSheet.cell(row=x+1,column=1,value=self.cheaters[x][0])
			cheatSheet.cell(row=x+1,column=2,value=self.cheaters[x][1])
			cheatSheet.cell(row=x+1,column=3,value=self.cheaters[x][2])
			cheatSheet.cell(row=x+1,column=4,value=self.cheaters[x][3])
			cheatSheet.cell(row=x+1,column=5,value=self.cheaters[x][4])

	def findSuspiciousPeople(self):
		for x in range(len(self.contestants)):
			p1 = {self.contestants[x][1],self.contestants[x][4]}
			for y in range(x+1,len(self.contestants)):
				p2 = {self.contestants[y][1],self.contestants[y][4]}
				if len(p1.intersection(p2)) == 2:
					self.suspiciousPeeps.append(self.contestants[x])
					self.suspiciousPeeps.append(self.contestants[y])
					break

	def deleteSuspiciousPerson(self):
		x = 0
		while x < (len(self.suspiciousPeeps)):
			answer = messagebox.askyesno("Question",'{}\n{}\nDelete from contestant list?'.format(formatPerson(self.suspiciousPeeps[x]),formatPerson(self.suspiciousPeeps[x+1])))				
			if answer == True:					
				if self.suspiciousPeeps[x] in self.contestants:
					self.contestants.remove(self.suspiciousPeeps[x])
				self.cheaters.append(self.suspiciousPeeps[x])
				if self.suspiciousPeeps[x+1] in self.contestants:
					self.contestants.remove(self.suspiciousPeeps[x+1])
				self.cheaters.append(self.suspiciousPeeps[x+1])				
				x += 2
			else:
				x += 2

	def publishContestants(self):
		for x in range(len(self.contestants)):
			contestantSheet.cell(row=x+1,column=1,value=self.contestants[x][0])
			contestantSheet.cell(row=x+1,column=2,value=self.contestants[x][1])
			contestantSheet.cell(row=x+1,column=3,value=self.contestants[x][2])
			contestantSheet.cell(row=x+1,column=4,value=self.contestants[x][3])
			contestantSheet.cell(row=x+1,column=5,value=self.contestants[x][4])

def formatPerson(person):
		return '{} {} {} {} {}'.format(person[0],person[1],person[4],person[3],person[2])

class Winners:

	def __init__(self,contestantList,numberofWinners):
		self.contestantList = contestantList
		self.winners = []
		self.numberofWinners = numberofWinners
		self.findWinners()
		self.publishWinners()

	def findWinners(self):
		for x in range(self.numberofWinners):
			winner = random.choice(self.contestantList)
			self.winners.append(winner)
			self.contestantList.remove(winner)	

	def publishWinners(self):
		sorted(self.winners,key=itemgetter(1))
		for x in range(len(self.winners)):
			winSheet.cell(row=x+1,column=1,value=self.winners[x][0])
			winSheet.cell(row=x+1,column=2,value=self.winners[x][1])
			winSheet.cell(row=x+1,column=3,value=self.winners[x][2])
			winSheet.cell(row=x+1,column=4,value=self.winners[x][3])
			winSheet.cell(row=x+1,column=5,value=self.winners[x][4])
			winSheet.cell(row=x+1,column=6,value=zips[self.winners[x][3][:3]])

contest = Contestants()
wb.save(filename)
#batchgeo for heat map



